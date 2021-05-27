# Creating a Python application program to analyze, update, and calculate .xlsx (Excel) data.

# Import the openpyxl pip package and bar chart and reference package
import openpyxl as xl
from openpyxl.chart import BarChart, Reference


# Create a function so the code can be used in many applications
# Note 'transactions.xlsx' is the argument passed to 'filename' in this use-case
def process_excel_file(filename):

    # Load excel workbook and create object wb
    wb = xl.load_workbook(filename)

    # Specify page number you wish to load
    sheet = wb['Sheet1']

    # Find out how many rows are on the sheet using max_row
    # print(sheet.max_row)

    # Specify the coordinate of the cell on the page
    # cell = sheet['a1']
    # sheet.cell(1, 1)

    # Print the cell value
    # print(cell.value)

    # Create a for loop to iterate over all rows and gather the price column (adding 1)
    # Then multiply it to create new price
    for row in range(2, sheet.max_row + 1):
        price_cell = sheet.cell(row, 3)
        new_price = price_cell.value * 0.9

        # Create a new column for the new prices just gathered
        new_price_cell = sheet.cell(row, 4)
        new_price_cell.value = new_price

    # Reference class to select a range of values (prices) for bar chart
    price_values = Reference(sheet,
              min_row=2,
              max_row=sheet.max_row,
              min_col=4,
              max_col=4)

    # Create instance of BarChart class and add data to it via new column
    chart_figures = BarChart()
    chart_figures.add_data(price_values)
    sheet.add_chart(chart_figures, 'e2')

    # Save to a new excel file for the new price column - in-case of a bug.
    # In this case, I saved 'filename' as 'transactions2.xlsx'
    wb.save(filename)
